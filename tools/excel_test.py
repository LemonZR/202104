import pprint

import openpyxl
from openpyxl.styles import fills, Font, Border, Side, PatternFill, colors, Alignment

# 读取 excel a
wb = openpyxl.load_workbook('a.xlsx')
sheets = wb.sheetnames

# 以下两个在这种场景等价
sheet1 = wb['Sheet1']  # 获取Sheet1
ws = wb.active  # 获取当前活动的sheet

print(sheets)
print(list(sheet1.rows))
print(ws)

data = list(sheet1.rows)
pprint.pprint(type(data[1][1]))


def unmerge_cell():
    ws.unmerge_cells('A1:B9')


def set_value():
    # 1 通过 单元格编号赋值
    A1 = sheet1['A1']
    A1.value = '姓名'

    # 通过sheet坐标赋值
    sheet1.cell(1, 2, '年龄')
    ws.cell(1, 3, '编号')


def insert_row():
    ws.insert_rows(idx=1, amount=2)


def merger():
    ws.merge_cells('A11:B11')
    a = ws.merged_cells
    print(a)


def patternFill():
    fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='000000')
    ws.cell(1, 1).fill = fill
    ws.cell(1, 2).fill = fills.GradientFill(stop=['FF0000', '0000FF'])
    ws.cell(1, 3).fill = fills.GradientFill(stop=['66ccff', '66ccff'])


patternFill()
wb.save('a.xlsx')
