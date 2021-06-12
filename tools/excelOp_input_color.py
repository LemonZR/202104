# coding=utf-8

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
import os, sys
from typing import Union

'''
import openpyxl
# 创建一个工作簿
wb = openpyxl.Workbook()
# 创建一个test_case的sheet表单
wb.create_sheet('test_case')
# 保存为一个xlsx格式的文件
wb.save('cases.xlsx')
# 读取excel中的数据
# 第一步：打开工作簿
wb = openpyxl.load_workbook('cases.xlsx')
# 第二步：选取表单
sh = wb['test_case']
# 第三步：读取数据
# 参数 row:行  column：列
ce = sh.cell(row = 1,column = 1)   # 读取第一行，第一列的数据
print(ce.value)
# 按行读取数据 list(sh.rows)
print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据

for cases in list(sh.rows)[1:]:
    case_id =  cases[0].value
    case_excepted = cases[1].value
    case_data = cases[2].value
    print(case_excepted,case_data)
# 关闭工作薄
wb.close()
'''


def write_xlsx(filename: str, data: list[Union[list[tuple], tuple[tuple]]], edit=False, sheet_name='newSheet') -> None:
    if edit:
        try:
            workbook = openpyxl.load_workbook(filename)
            is_sheet = workbook[sheet_name]
            workbook.remove(is_sheet)
            print(f'remove sheet[{sheet_name}] in workbook[{filename}],I\'ll create new one ')
        except FileNotFoundError as fe:
            print(f'{fe},\nI\'ll create a new workbook named {filename}')
            workbook = openpyxl.Workbook()
        except KeyError as ke:
            print(f'{ke},\nI\'ll create a new sheet named [{sheet_name}] in workbook [{filename}]')
        except Exception as e:
            sys.exit(e)
    elif os.path.exists(filename):
        sys.exit(
            f'Workbook [{filename}] already exists,you can\'t write it.\nPlease set [edit=True],then try again!')
    else:
        workbook = openpyxl.Workbook()
        print(f'Create a new workbook named[{filename}]')
    new_sheet = workbook.create_sheet(sheet_name, index=0)

    for row_id, row_data in enumerate(data, start=1):
        # data :[[(x1,color1),(x2,color2)],[(y1,color3),(y2,color4),(y3,color3)],[(z1,color5),(z2,color6)]]
        # row_data: [(x1,color1),(x2,color2)]
        if row_id == 1:
            font = Font(b=True)
        else:
            font = Font(b=False)
        side = Side(style='thin', color='000000')
        cell_border = Border(left=side, right=side, top=side, bottom=side)
        for col_id, cell_data in enumerate(row_data, start=1):
            # cell_data : (x1,color1)
            fgColor = cell_data[-1] if cell_data and cell_data[-1] else 'FFFFFF'
            cell_fill = PatternFill(patternType='solid', fgColor=fgColor)
            new_sheet.cell(row_id, col_id).fill = cell_fill
            new_sheet.cell(row_id, col_id).border = cell_border
            new_sheet.cell(row_id, col_id).font = font

            new_sheet.cell(row_id, col_id, cell_data[0])
    workbook.save(filename)
    workbook.close()


def read_xlsx(filename, sheet_name='newSheet') -> list[list]:
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]

    except FileNotFoundError as fe:

        sys.exit(fe)

    except KeyError as ke:
        sys.exit(ke)

    result_list = []
    for row_data in sheet.rows:
        row_list = []
        for cell in row_data:
            row_list.append(cell.value)
        result_list.append(row_list)
    workbook.close()
    return result_list


if __name__ == '__main__':
    basedir = os.path.dirname(os.path.dirname(__file__)) + '\\dustbin'
    fileName = f'{basedir}\\test.xlsx'
    # write_xlsx(fileName,
    #            data=[[('head',None), ('head',None), ('head',None)], [('asdc', 'FF6347'), ('asd',None)], [('asdas', 'FF6347')]],
    #            edit=True, sheet_name='Sheet')
    # result = read_xlsx(fileName, sheet_name='Sheet')
    #
    # for i in result:
    #     print(i)
    # print(write_xlsx.__annotations__)
    # print(read_xlsx.__annotations__)
    cs = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', 'A', 'B', 'C', 'D', 'E', 'F']

    import random
    #
    # a = random.choice(cs)
    # b = random.choice(cs)
    # c = random.choice(cs)
    # d = random.choice(cs)
    # e = random.choice(cs)
    # f = random.choice(cs)
    #
    # data = []
    # for i in range(100):
    #     cc = []
    #     for j in range(150):
    #         g = random.choices(cs,[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16], k=6)
    #         color = ''.join(g)
    #         print(g)
    #         cc += [('', color)]
    #     data.append(cc)
    #
    # write_xlsx(fileName, data, edit=True)
    i = 0
    for j in range(10000000):
        g = random.choices(cs, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16], k=6)
        i +=1
        if ''.join(g) =='FFFFFF':
            print(i)
            break