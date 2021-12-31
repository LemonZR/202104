import datetime
from Reporter import Reporter
import openpyxl
import sys

import config

init_config = config.init_config


class MkDayReporter(Reporter):

    def __init__(self, check_date=None):

        default_trade_date = (datetime.date.today() - datetime.timedelta(days=3)).strftime("%Y%m%d")
        check_date = check_date if check_date else default_trade_date
        # 初始化参数
        conf = init_config(check_date)
        super(MkDayReporter, self).__init__(**conf['mk_day'])

        self.result_column_index = self.info_column_index + 3  # 因为这里设置了[处理结果列]的索引，所以不需要result_column_pattern了

    def analyze(self):
        return super().analyze()

    def export_excel(self, report_path='', template_path=None):
        if not template_path:
            template_path = report_path
        # 得先执行 analyze 再获取report
        if self.report:
            try:
                workbook = openpyxl.load_workbook(rf'{template_path}')
                sheet = workbook['Sheet1']
            except FileNotFoundError as fe:
                sys.exit(fe)
            sheet['C4'] = self.check_date
            sheet['C3'] = self.check_date
            sheet['D4'] = self.report['am']['总量']
            sheet['F4'] = self.report['am'][0] + self.report['am'][5]
            sheet['E4'] = self.report['am'][1]
            sheet['G4'] = self.report['am'][2]
            sheet['H4'] = self.report['am'][3]
            sheet['I4'] = self.report['am'][4]
            # sheet['K4'] = self.report['am']['剩余数量']
            sheet['J4'] = self.report['am']['已完成']

            sheet['D3'] = self.report['mk']['总量']
            sheet['F3'] = self.report['mk'][0] + self.report['am'][5]
            sheet['E3'] = self.report['mk'][1]
            sheet['G3'] = self.report['mk'][2]
            sheet['H3'] = self.report['mk'][3]
            sheet['I3'] = self.report['mk'][4]
            # sheet['K3'] = self.report['mk']['剩余数量']
            sheet['J3'] = self.report['mk']['已完成']
            workbook.save(report_path)


class MkMonReporter(Reporter):
    def __init__(self, check_date=None):
        default_trade_date = datetime.date.today().strftime("%Y") + str(datetime.date.today().month - 1)
        check_date = check_date if check_date else default_trade_date
        conf = init_config(check_date)
        super(MkMonReporter, self).__init__(**conf['mk_month'])

    def analyze(self):
        return super().analyze()

    def export_excel(self, report_path='', template_path=None):
        if not template_path:
            template_path = report_path
        # 得先执行 analyze 再获取report
        if self.report:
            try:
                workbook = openpyxl.load_workbook(rf'{template_path}')
                sheet = workbook['Sheet1']
            except FileNotFoundError as fe:
                sys.exit(fe)
            sheet['C6'] = self.check_date
            sheet['D6'] = self.report['am']['总量']
            sheet['F6'] = self.report['am'][0] + self.report['am'][5]
            sheet['E6'] = self.report['am'][1]
            sheet['G6'] = self.report['am'][2]
            sheet['H6'] = self.report['am'][3]
            sheet['I6'] = self.report['am'][4]
            # sheet['K6'] = self.report['am']['剩余数量']
            sheet['J6'] = self.report['am']['已完成']

            sheet['C5'] = self.check_date
            sheet['D5'] = self.report['mk']['总量']
            sheet['F5'] = self.report['mk'][0] + self.report['am'][5]
            sheet['E5'] = self.report['mk'][1]
            sheet['G5'] = self.report['mk'][2]
            sheet['H5'] = self.report['mk'][3]
            sheet['I5'] = self.report['mk'][4]
            # sheet['K5'] = self.report['mk']['剩余数量']
            sheet['J5'] = self.report['mk']['已完成']
            workbook.save(report_path)


class DisDayReporter(Reporter):
    def __init__(self, check_date=None):
        default_trade_date = (datetime.date.today() - datetime.timedelta(days=6)).strftime("%Y%m%d")
        check_date = check_date if check_date else default_trade_date
        conf = init_config(check_date)
        super(DisDayReporter, self).__init__(**conf['dis_day'])

    def analyze(self):
        return super().analyze()

    def export_excel(self, report_path='', template_path=None):
        if not template_path:
            template_path = report_path
        # 得先执行 analyze 再获取report
        if self.report:
            try:
                workbook = openpyxl.load_workbook(rf'{template_path}')
                sheet = workbook['Sheet1']
            except FileNotFoundError as fe:
                sys.exit(fe)

            # 专题
            sheet['C7'] = self.check_date
            sheet['D7'] = self.report[1]['总量']
            sheet['F7'] = self.report[1][0] + self.report[1][5]
            sheet['E7'] = self.report[1][1]
            sheet['G7'] = self.report[1][2]
            sheet['H7'] = self.report[1][3]
            sheet['I7'] = self.report[1][4]
            # sheet['K6'] = self.report['am']['剩余数量']
            sheet['J7'] = self.report[1]['已完成']
            # 应用依赖
            sheet['C9'] = self.check_date
            sheet['D9'] = self.report[0]['总量']
            sheet['F9'] = self.report[0][0] + self.report[0][5]
            sheet['E9'] = self.report[0][1]
            sheet['G9'] = self.report[0][2]
            sheet['H9'] = self.report[0][3]
            sheet['I9'] = self.report[0][4]
            # sheet['K9'] = self.report['mk']['剩余数量']
            sheet['J9'] = self.report[0]['已完成']
            # top100报表
            sheet['C8'] = self.check_date
            sheet['D8'] = self.report[2]['总量']
            sheet['F8'] = self.report[2][0] + self.report[2][5]
            sheet['E8'] = self.report[2][1]
            sheet['G8'] = self.report[2][2]
            sheet['H8'] = self.report[2][3]
            sheet['I8'] = self.report[2][4]
            # sheet['K5'] = self.report['mk']['剩余数量']
            sheet['J8'] = self.report[2]['已完成']
            workbook.save(report_path)


class DisMonthReporter(Reporter):
    def __init__(self, check_date=None):
        default_trade_date = datetime.date.today().strftime("%Y") + str(datetime.date.today().month - 1)
        check_date = check_date if check_date else default_trade_date
        conf = init_config(check_date)
        super(DisMonthReporter, self).__init__(**conf['dis_month'])

    def analyze(self):
        return super().analyze()

    def export_excel(self, report_path='', template_path=None):
        pass


if __name__ == '__main__':
    # 日报输出位置
    reportpath = r'C:\Users\zwx1000092\Desktop\集中化数据核对日报.xlsx'
    #
    # mk(am)日模型
    md = MkDayReporter()
    md.analyze()
    md.export_excel(reportpath)  # 不写入日报就注释掉这行
    #
    #
    # mk(am)月模型
    mm = MkMonReporter('202111')
    mm.analyze()
    mm.export_excel(reportpath)

    # dis 日模型
    # dd = DisDayReporter('20211217')
    # dd.analyze()
    # dd.export_excel(reportpath)
